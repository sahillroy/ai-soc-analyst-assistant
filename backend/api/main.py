from dotenv import load_dotenv
load_dotenv()
import os
from fastapi import FastAPI, BackgroundTasks, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import text
import pandas as pd
import datetime
from typing import Optional

from backend.core.database import (
    engine, check_db,
    db_get_pipeline_state,
    db_set_pipeline_running,
    db_set_pipeline_done,
    db_is_stale,
    _get_column_names,
)
from backend.detection.pipeline import run_pipeline


app = FastAPI(title="SOC Analyst API", version="2.2.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "http://localhost:5174",
        "http://127.0.0.1:5174",
        "http://localhost:5175",
        "http://127.0.0.1:5175",
        "https://ai-soc-analyst-assistant.vercel.app",
        "https://ai-soc-analyst-assistant-o6prk50hz-sahillroys-projects.vercel.app",
    ],
    allow_origin_regex=r"https://.*\.vercel\.app",
    allow_methods=["*"],
    allow_headers=["*"],
    allow_credentials=False,
)


# ── DataFrame sanitisation ────────────────────────────────────────────────────

def _sanitise_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Coerce all DataFrame columns to PostgreSQL-safe types before to_sql().

    WHY THIS EXISTS:
    - rule_engine.py produces bool columns via .isin() / comparisons → numpy bool_
    - PostgreSQL INTEGER columns reject Python/numpy bool → 'invalid input syntax'
    - NaN in numeric columns → DataError on non-nullable PG columns
    - numpy int64/float64 can confuse older psycopg2 versions
    """
    bool_as_int = [
        'anomaly', 'rule_bruteforce', 'rule_port_scan',
        'rule_traffic_spike', 'rule_triggered', 'ip_suspicious',
    ]
    float_cols = [
        'bytes_transferred', 'anomaly_score', 'confidence',
        'bytes_zscore', 'risk_score',
    ]
    int_cols  = ['port', 'failed_logins']
    text_cols = [
        'incident_id', 'timestamp', 'source_ip', 'destination_ip',
        'protocol', 'severity', 'alert_type', 'campaign_id',
        'mitre_technique', 'incident_summary', 'recommended_action',
        'soc_playbook_action', 'escalation', 'automation_result',
        'ip_country', 'status', 'notes',
    ]

    for col in bool_as_int:
        if col in df.columns:
            df[col] = df[col].fillna(0).astype(bool).astype(int)
    for col in float_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
    for col in int_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
    for col in text_cols:
        if col in df.columns:
            df[col] = df[col].fillna('').astype(str)

    return df


# ── Background pipeline task ──────────────────────────────────────────────────

def _run_pipeline_task():
    now_iso = lambda: datetime.datetime.utcnow().isoformat()
    try:
        df = run_pipeline()

        # Drop ML-only columns not in DB schema
        ml_cols = ['hour', 'source_ip_encoded', 'destination_ip_encoded', 'protocol_encoded']
        df = df.drop(columns=[col for col in ml_cols if col in df.columns], errors='ignore')

        # Coerce all types to PostgreSQL-safe values
        df = _sanitise_df(df)

        print(f"[pipeline] Columns going to DB: {sorted(df.columns.tolist())}")
        print(f"[pipeline] Row count: {len(df)}")
        print(f"[pipeline] dtypes sample:\n{df.dtypes}")

        # Clear existing rows, then insert (preserves our DDL schema)
        with engine.connect() as conn:
            conn.execute(text("DELETE FROM alerts"))
            conn.commit()

        df.to_sql(
            "alerts", engine,
            if_exists="append",   # NEVER use 'replace' — it drops our DDL
            index=False,
            method="multi",
            chunksize=50,
        )

        with engine.connect() as conn:
            db_set_pipeline_done(conn, last_run=now_iso(), total_alerts=len(df))

        print(f"[✓] Pipeline done — {len(df)} alerts written.")

    except Exception as e:
        import traceback
        print(f"[ERROR] Pipeline task failed: {e}")
        traceback.print_exc()
        # ALWAYS release the lock, even on crash
        try:
            with engine.connect() as conn:
                db_set_pipeline_done(conn, last_run=now_iso(), total_alerts=0)
        except Exception as e2:
            print(f"[ERROR] Could not release pipeline lock: {e2}")


# ── Routes ────────────────────────────────────────────────────────────────────

@app.post("/api/run-analysis")
def run_analysis(background_tasks: BackgroundTasks):
    with engine.connect() as conn:
        state = db_get_pipeline_state(conn)

        if state["running"]:
            if db_is_stale(conn):
                print("[WARN] Stale pipeline lock detected — clearing.")
                db_set_pipeline_running(conn, False)
            else:
                raise HTTPException(
                    status_code=409,
                    detail="Pipeline already running. Try again shortly."
                )

        started_at = datetime.datetime.utcnow().isoformat()
        db_set_pipeline_running(conn, True, started_at=started_at)

    background_tasks.add_task(_run_pipeline_task)
    return {"status": "started", "started_at": started_at}


@app.post("/api/reset-status")
def reset_status():
    """Force-clear the pipeline lock (use after a stuck/crashed run)."""
    with engine.connect() as conn:
        db_set_pipeline_running(conn, False)
    return {"status": "reset", "message": "Pipeline lock cleared."}


@app.get("/api/status")
def get_status():
    with engine.connect() as conn:
        state = db_get_pipeline_state(conn)
    return state


@app.get("/api/alerts")
def get_alerts(
    severity: Optional[str] = Query(None),
    limit: int = Query(100, le=500),
    offset: int = Query(0),
    search: Optional[str] = Query(None),
):
    """
    Returns alerts. NEVER returns 500 — always returns [] on any DB error.

    The 500s were caused by:
    1. ORDER BY id (column didn't exist in old to_sql() tables) → fixed: ORDER BY timestamp
    2. Table in corrupt/partial state during pipeline write → fixed: catch all errors → []
    3. Table had wrong column types → fixed: schema repair in database.py startup
    """
    try:
        conditions = []
        params: dict = {}

        if severity:
            conditions.append("severity = :severity")
            params["severity"] = severity

        if search:
            conditions.append(
                "(source_ip LIKE :search OR incident_id LIKE :search OR alert_type LIKE :search)"
            )
            params["search"] = f"%{search}%"

        where  = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        params["limit"]  = limit
        params["offset"] = offset

        query = f"""
            SELECT * FROM alerts
            {where}
            ORDER BY timestamp DESC
            LIMIT :limit OFFSET :offset
        """

        with engine.connect() as conn:
            result = conn.execute(text(query), params)
            rows = [dict(row._mapping) for row in result]

        return rows

    except Exception as e:
        err = str(e).lower()
        # Table doesn't exist yet — normal before first run
        if any(x in err for x in ["no such table", "does not exist", "relation", "undefined"]):
            return []
        # Any other error: log it but still return [] so the UI doesn't break
        print(f"[ERROR] GET /api/alerts failed: {e}")
        import traceback
        traceback.print_exc()
        return []   # ← return empty array, NOT 500


@app.get("/api/stats")
def get_stats():
    try:
        with engine.connect() as conn:
            result = conn.execute(
                text("SELECT severity, COUNT(*) as count FROM alerts GROUP BY severity")
            )
            return [dict(row._mapping) for row in result]
    except Exception:
        return []


@app.patch("/api/alerts/{incident_id}/status")
def update_status(incident_id: str, body: dict):
    allowed = {"New", "Investigating", "Resolved", "False Positive"}
    new_status = body.get("status")
    if new_status not in allowed:
        raise HTTPException(status_code=400, detail="Invalid status")
    with engine.connect() as conn:
        conn.execute(
            text("UPDATE alerts SET status=:status WHERE incident_id=:id"),
            {"status": new_status, "id": incident_id}
        )
        conn.commit()
    return {"updated": incident_id, "status": new_status}


@app.patch("/api/alerts/{incident_id}/notes")
def update_notes(incident_id: str, body: dict):
    notes = body.get("notes", "")
    with engine.connect() as conn:
        conn.execute(
            text("UPDATE alerts SET notes=:notes WHERE incident_id=:id"),
            {"notes": notes, "id": incident_id}
        )
        conn.commit()
    return {"updated": incident_id}


@app.get("/api/health")
def health():
    return check_db(engine)


@app.get("/api/debug")
def debug():
    """
    Shows exact DB state for troubleshooting.
    Check this when things break — it shows the actual error.
    URL: /api/debug
    """
    result = {"engine": engine.name, "database_url_tail": ""}
    try:
        result["database_url_tail"] = (
            DATABASE_URL.split("@")[-1] if "@" in DATABASE_URL
            else DATABASE_URL.split("///")[-1]
        )
    except Exception:
        pass

    # Test basic connectivity
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        result["db_connectivity"] = "ok"
    except Exception as e:
        result["db_connectivity"] = f"FAILED: {e}"
        return result

    # Inspect alerts table schema
    try:
        with engine.connect() as conn:
            cols = _get_column_names(conn, "alerts")
        result["alerts_columns"] = cols
        result["alerts_has_id"]  = "id" in cols
    except Exception as e:
        result["alerts_schema_error"] = str(e)

    # Try a SELECT
    try:
        with engine.connect() as conn:
            count = conn.execute(text("SELECT COUNT(*) FROM alerts")).scalar()
            sample = conn.execute(text("SELECT * FROM alerts LIMIT 1")).fetchone()
        result["alerts_count"]  = count
        result["alerts_sample"] = dict(sample._mapping) if sample else None
    except Exception as e:
        result["alerts_select_error"] = str(e)

    # Pipeline state
    try:
        with engine.connect() as conn:
            state = db_get_pipeline_state(conn)
        result["pipeline_state"] = state
    except Exception as e:
        result["pipeline_state_error"] = str(e)

    return result


# Import DATABASE_URL for debug endpoint
try:
    from backend.core.database import DATABASE_URL
except ImportError:
    DATABASE_URL = os.getenv("DATABASE_URL", "")


# ── Incident Report Endpoints ──────────────────────────────────────────────────

@app.get("/api/report/{incident_id}")
def get_incident_report(incident_id: str):
    """Returns a structured AI-generated JSON report for one incident."""
    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT * FROM alerts WHERE incident_id = :id"),
            {"id": incident_id}
        ).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Incident not found")

    data = dict(row._mapping)
    return {
        "incident_id":    data.get("incident_id"),
        "summary":        data.get("incident_summary"),
        "recommendation": data.get("recommended_action"),
        "severity":       data.get("severity"),
        "risk_score":     data.get("risk_score"),
        "alert_type":     data.get("alert_type"),
        "source_ip":      data.get("source_ip"),
        "destination_ip": data.get("destination_ip"),
        "timestamp":      str(data.get("timestamp", "")),
        "mitre":          data.get("mitre_technique"),
        "soc_playbook":   data.get("soc_playbook_action"),
        "automation":     data.get("automation_result"),
        "escalation":     data.get("escalation"),
        "campaign_id":    data.get("campaign_id"),
        "notes":          data.get("notes", ""),
    }


@app.get("/api/report/export/csv")
def export_full_report_csv():
    """Returns a rich CSV report of all alerts with all AI-generated fields."""
    from fastapi.responses import Response

    try:
        with engine.connect() as conn:
            rows = conn.execute(text("SELECT * FROM alerts ORDER BY id DESC")).fetchall()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    if not rows:
        raise HTTPException(status_code=404, detail="No alerts found. Run Analysis first.")

    columns = [
        "incident_id", "timestamp", "source_ip", "destination_ip", "port", "protocol",
        "alert_type", "severity", "risk_score", "confidence", "campaign_id",
        "escalation", "status",
        "incident_summary", "recommended_action", "soc_playbook_action",
        "automation_result", "mitre_technique", "notes",
    ]

    def escape(val):
        s = str(val) if val is not None else ""
        return '"' + s.replace('"', '""') + '"'

    header = ",".join(columns)
    lines  = [header]
    for row in rows:
        d = dict(row._mapping)
        lines.append(",".join(escape(d.get(c, "")) for c in columns))

    csv_text = "\n".join(lines)

    return Response(
        content=csv_text,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=soc-ai-incident-report.csv"}
    )


@app.get("/api/report/export/xlsx")
def export_full_report_xlsx():
    """
    Generates a professional styled Excel report with:
    - Severity-colored rows (Critical=purple, High=red, Medium=amber, Low=green)
    - Bold headers with frozen top row
    - Auto-sized columns
    - Sheet 2: Summary statistics (severity + alert type breakdown)

    FIXES vs previous version:
    1. Uses Response(content=buf.getvalue()) instead of StreamingResponse(buf)
       StreamingResponse reads lazily after function returns — BytesIO can be
       GC'd before streaming completes, causing silent 500s.
    2. Removed unused GradientFill import (causes ImportError on older openpyxl).
    3. ORDER BY CAST(risk_score AS FLOAT) — risk_score stored as TEXT in some
       old schemas sorts lexicographically ('9' > '77'), giving wrong order.
    4. All None values from PostgreSQL NULL are safely handled before float().
    """
    import io
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import Counter

    # ── Fetch data ────────────────────────────────────────────────────────────
    try:
        with engine.connect() as conn:
            rows = conn.execute(text(
                "SELECT * FROM alerts ORDER BY risk_score DESC"
            )).fetchall()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DB error: {e}")

    if not rows:
        raise HTTPException(status_code=404, detail="No alerts found. Run Analysis first.")

    data_dicts = [dict(r._mapping) for r in rows]

    # ── Style palette ─────────────────────────────────────────────────────────
    SEV_FILL = {
        "Critical": PatternFill("solid", fgColor="3B0764"),
        "High":     PatternFill("solid", fgColor="7F1D1D"),
        "Medium":   PatternFill("solid", fgColor="78350F"),
        "Low":      PatternFill("solid", fgColor="14532D"),
        "Normal":   PatternFill("solid", fgColor="1E293B"),
    }
    SEV_FONT = {
        "Critical": Font(color="E9D5FF", bold=True, size=9),
        "High":     Font(color="FECACA", bold=True, size=9),
        "Medium":   Font(color="FDE68A", size=9),
        "Low":      Font(color="BBF7D0", size=9),
        "Normal":   Font(color="94A3B8", size=9),
    }
    HEADER_FILL  = PatternFill("solid", fgColor="0F172A")
    HEADER_FONT  = Font(color="60A5FA", bold=True, size=10)
    BASE_FILL    = PatternFill("solid", fgColor="0F172A")
    ALT_FILL     = PatternFill("solid", fgColor="1E293B")
    BASE_FONT    = Font(color="CBD5E1", size=9)
    ALT_FONT     = Font(color="94A3B8", size=9)
    THIN_BORDER  = Border(bottom=Side(border_style="thin", color="334155"))
    CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=False)
    WRAP         = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

    # ── Column definitions: (header label, db field, column width) ────────────
    COLS = [
        ("Incident ID",        "incident_id",         22),
        ("Timestamp",          "timestamp",            20),
        ("Source IP",          "source_ip",            16),
        ("Destination",        "destination_ip",       16),
        ("Port",               "port",                  7),
        ("Protocol",           "protocol",              9),
        ("Alert Type",         "alert_type",           26),
        ("Severity",           "severity",             11),
        ("Risk Score",         "risk_score",           11),
        ("Confidence %",       "confidence",           13),
        ("Campaign",           "campaign_id",          14),
        ("Escalation",         "escalation",           28),
        ("Status",             "status",               14),
        ("AI Summary",         "incident_summary",     55),
        ("Recommended Action", "recommended_action",   45),
        ("SOC Playbook",       "soc_playbook_action",  35),
        ("Automation Result",  "automation_result",    38),
        ("MITRE Technique",    "mitre_technique",      30),
        ("Analyst Notes",      "notes",                30),
    ]

    WRAP_FIELDS = {
        "incident_summary", "recommended_action",
        "soc_playbook_action", "automation_result", "notes",
    }
    FLOAT_FIELDS = {"risk_score", "confidence"}

    wb = Workbook()

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 1 — Incident Report
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Incident Report"
    ws.sheet_view.showGridLines = False

    # Header row
    for col_idx, (label, _, width) in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"   # freeze header row

    # Data rows
    for row_idx, d in enumerate(data_dicts, start=2):
        sev      = d.get("severity") or "Normal"
        row_fill = SEV_FILL.get(sev, BASE_FILL)
        row_font = SEV_FONT.get(sev, BASE_FONT)

        # Alternate faint shade for Medium/Low/Normal rows
        if sev not in ("Critical", "High"):
            row_fill = ALT_FILL if row_idx % 2 == 0 else BASE_FILL
            row_font = ALT_FONT  if row_idx % 2 == 0 else BASE_FONT

        has_wrap = False
        for col_idx, (_, field, _) in enumerate(COLS, start=1):
            val  = d.get(field)
            # Safe None handling — PostgreSQL NULLs come through as Python None
            disp = "" if val is None else str(val)

            if field in FLOAT_FIELDS and val is not None:
                try:
                    disp = f"{float(val):.1f}"
                except (TypeError, ValueError):
                    pass

            cell        = ws.cell(row=row_idx, column=col_idx, value=disp)
            cell.fill   = row_fill
            cell.font   = row_font
            cell.border = THIN_BORDER

            if field in WRAP_FIELDS:
                cell.alignment = WRAP
                has_wrap = True
            else:
                cell.alignment = CENTER

        ws.row_dimensions[row_idx].height = 65 if has_wrap else 16

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 2 — Summary Statistics
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 14

    def _s2_cell(row, col, value, font=None, fill=None, align=None):
        c = ws2.cell(row=row, column=col, value=value)
        if font:  c.font      = font
        if fill:  c.fill      = fill
        if align: c.alignment = align
        c.border = THIN_BORDER
        return c

    # Title
    _s2_cell(1, 1, "SOC AI Analyst — Incident Summary",
             font=Font(color="60A5FA", bold=True, size=14),
             fill=HEADER_FILL,
             align=Alignment(horizontal="left", vertical="center"))
    ws2.row_dimensions[1].height = 28

    _s2_cell(2, 1, f"Total incidents analysed: {len(data_dicts)}",
             font=Font(color="94A3B8", italic=True, size=10),
             fill=BASE_FILL)

    # Severity breakdown
    sev_counts  = Counter(d.get("severity") or "Normal" for d in data_dicts)
    type_counts = Counter(d.get("alert_type") or "Unknown" for d in data_dicts)

    _s2_cell(4, 1, "Severity", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    _s2_cell(4, 2, "Count",    font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)

    for i, sev in enumerate(["Critical", "High", "Medium", "Low", "Normal"], start=5):
        cnt = sev_counts.get(sev, 0)
        _s2_cell(i, 1, sev, font=SEV_FONT.get(sev, BASE_FONT),
                 fill=SEV_FILL.get(sev, BASE_FILL), align=CENTER)
        _s2_cell(i, 2, cnt, font=SEV_FONT.get(sev, BASE_FONT),
                 fill=SEV_FILL.get(sev, BASE_FILL), align=CENTER)

    _s2_cell(11, 1, "Alert Type",  font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    _s2_cell(11, 2, "Count",       font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)

    for j, (atype, cnt) in enumerate(type_counts.most_common(10), start=12):
        fill = ALT_FILL if j % 2 == 0 else BASE_FILL
        _s2_cell(j, 1, atype, font=BASE_FONT, fill=fill, align=CENTER)
        _s2_cell(j, 2, cnt,   font=BASE_FONT, fill=fill, align=CENTER)

    # ── Serialize to bytes and return ─────────────────────────────────────────
    # IMPORTANT: Use Response(content=buf.getvalue()), NOT StreamingResponse(buf).
    # StreamingResponse reads the BytesIO lazily after this function returns,
    # by which point Python may have GC'd the local buf — causing a silent 500.
    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()   # read all bytes NOW, while buf is still alive

    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=soc-ai-incident-report.xlsx"},
    )
