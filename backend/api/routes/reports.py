"""
routes/reports.py — Incident reporting & export endpoints

Handles:
  GET /api/report/{incident_id}  — single incident JSON report
  GET /api/report/export/xlsx    — full Excel export
  GET /api/report/export/csv     — full CSV export
"""
import io
import json

import pandas as pd
from fastapi import APIRouter, HTTPException
from fastapi.responses import Response, StreamingResponse
from sqlalchemy import text

from backend.core.database import engine

router = APIRouter()


@router.get("/report/export/xlsx")
def export_report_xlsx():
    """Download a styled Excel report of all alerts."""
    try:
        with engine.connect() as conn:
            rows = conn.execute(text("SELECT * FROM alerts ORDER BY risk_score DESC")).fetchall()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    if not rows:
        raise HTTPException(status_code=404, detail="No alerts found. Run Analysis first.")

    from datetime import datetime
    data_dicts = [dict(r._mapping) for r in rows]
    generated_at = datetime.utcnow().strftime('%B %d, %Y  %H:%M UTC')

    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # ── COLOR PALETTE (matches Sentinel Core design system) ───────────────────────
    C = {
        'bg_dark':      '0D1526',
        'bg_card':      '0F1E35',
        'bg_header':    '0A1628',
        'bg_alt':       '111F36',
        'accent_blue':  '3B82F6',
        'accent_purple':'8B5CF6',
        'accent_red':   'EF4444',
        'accent_amber': 'F59E0B',
        'accent_green': '10B981',
        'text_primary': 'F1F5F9',
        'text_secondary':'94A3B8',
        'text_muted':   '475569',
        'border_light': '1E293B',
        'border_blue':  '1E3A5F',
        'white':        'FFFFFF',
    }
    
    SEV_COLORS = {
        'Critical': ('8B5CF6', '3B0764', 'E9D5FF'),
        'High':     ('EF4444', '7F1D1D', 'FECACA'),
        'Medium':   ('F59E0B', '78350F', 'FDE68A'),
        'Low':      ('10B981', '14532D', 'BBF7D0'),
    }
    
    def _fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)
    
    def _font(color=None, bold=False, size=10, name='Calibri', italic=False):
        return Font(name=name, size=size, bold=bold, italic=italic,
                    color=color or C['text_primary'])
    
    def _align(h='left', v='center', wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    
    def _border_bottom(color=None):
        b = Side(border_style='thin', color=color or C['border_light'])
        n = Side(border_style=None)
        return Border(bottom=b, left=n, right=n, top=n)
    
    def _border_full(color):
        s = Side(border_style='thin', color=color)
        return Border(left=s, right=s, top=s, bottom=s)
    
    # ── COLUMN DEFINITIONS ─────────────────────────────────────────────────────────
    COLS = [
        ('INCIDENT ID',     20, 'left'),
        ('TIMESTAMP',       18, 'center'),
        ('SOURCE IP',       16, 'center'),
        ('DEST IP',         14, 'center'),
        ('PORT',             7, 'center'),
        ('ALERT TYPE',      24, 'left'),
        ('SEVERITY',        12, 'center'),
        ('RISK SCORE',      10, 'center'),
        ('CONFIDENCE',      12, 'center'),
        ('CAMPAIGN',        14, 'center'),
        ('ESCALATION',      16, 'center'),
        ('MITRE ATT&CK',    32, 'left'),
        ('AI SUMMARY',      52, 'left'),
        ('RECOMMENDATION',  48, 'left'),
    ]
    
    wb = Workbook()
    
    # ══════════════════════════════════════════════════════════════════════════════
    # SHEET 1 — INCIDENT REPORT
    # ══════════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = 'Incident Report'
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False
    
    for i, (_, w, _) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    # Row 1: Title banner
    ws.row_dimensions[1].height = 44
    ws.merge_cells('A1:N1')
    c = ws['A1']
    c.value = '  SENTINEL CORE  ·  SOC INCIDENT REPORT'
    c.font  = _font(C['white'], bold=True, size=16)
    c.fill  = _fill(C['bg_dark'])
    c.alignment = _align('left', 'center')
    c.border = Border(left=Side(border_style='thick', color=C['accent_blue']))
    
    # Row 2: Meta / subtitle
    ws.row_dimensions[2].height = 22
    ws.merge_cells('A2:N2')
    c = ws['A2']
    c.value = f"  Generated: {generated_at}  ·  Total Incidents: {len(data_dicts)}"
    c.font  = _font(C['text_secondary'], size=9, italic=True)
    c.fill  = _fill(C['bg_dark'])
    c.alignment = _align('left', 'center')
    
    # Row 3: Spacer
    ws.row_dimensions[3].height = 8
    for col in range(1, 15):
        ws.cell(3, col).fill = _fill(C['bg_dark'])
    
    # Row 4: Column headers
    ws.row_dimensions[4].height = 30
    for i, (label, _, h_align) in enumerate(COLS, 1):
        c = ws.cell(4, i)
        c.value = label
        c.font  = _font(C['accent_blue'], bold=True, size=9)
        c.fill  = _fill(C['bg_header'])
        c.alignment = _align(h_align, 'center')
        c.border = Border(
            bottom=Side(border_style='medium', color=C['accent_blue']),
            left=Side(border_style=None), right=Side(border_style=None),
            top=Side(border_style=None),
        )
    
    ws.freeze_panes = 'A5'
    
    # Map DB field names to COLS index
    FIELD_MAP = [
        'incident_id', 'timestamp', 'source_ip', 'destination_ip', 'port',
        'alert_type', 'severity', 'risk_score', 'confidence', 'campaign_id',
        'escalation', 'mitre_technique', 'incident_summary', 'recommended_action',
    ]
    
    for row_i, d in enumerate(data_dicts):
        excel_row = row_i + 5
        is_alt    = (row_i % 2 == 1)
        row_bg    = C['bg_alt'] if is_alt else C['bg_card']
    
        sev        = str(d.get('severity') or '')
        sev_colors = SEV_COLORS.get(sev, (C['text_muted'], C['bg_card'], C['text_primary']))
    
        # Row height — based on longest text field
        summary_len = len(str(d.get('incident_summary') or ''))
        rec_len     = len(str(d.get('recommended_action') or ''))
        row_h = min(max(42, (max(summary_len, rec_len) // 80 + 1) * 14 + 10), 90)
        ws.row_dimensions[excel_row].height = row_h
    
        bb = _border_bottom()
    
        for col_i, field in enumerate(FIELD_MAP):
            c   = ws.cell(excel_row, col_i + 1)
            val = d.get(field)
            disp = '' if val is None else str(val)
    
            c.fill   = _fill(row_bg)
            c.border = bb
    
            # ── Per-column styling ─────────────────────────────────────────────
            if col_i == 0:   # Incident ID
                c.value = disp
                c.font  = _font(C['accent_blue'], bold=True, size=9, name='Consolas')
                c.alignment = _align('left', 'center')
    
            elif col_i == 1:  # Timestamp
                c.value = disp
                c.font  = _font(C['text_secondary'], size=9, name='Consolas')
                c.alignment = _align('center', 'center')
    
            elif col_i in (2, 3):  # Source / Dest IP
                c.value = disp
                c.font  = _font(C['text_primary'], size=9, name='Consolas')
                c.alignment = _align('center', 'center')
    
            elif col_i == 4:  # Port
                c.value = disp
                c.font  = _font(C['text_muted'], size=9)
                c.alignment = _align('center', 'center')
    
            elif col_i == 5:  # Alert Type
                c.value = disp
                c.font  = _font(C['text_primary'], size=9)
                c.alignment = _align('left', 'center')
    
            elif col_i == 6:  # Severity badge
                sev_bg, sev_dark, sev_light = sev_colors
                c.value = disp.upper()
                c.font  = _font(sev_light, bold=True, size=8)
                c.fill  = _fill(sev_dark)
                c.alignment = _align('center', 'center')
                badge_side = Side(border_style='thin', color=sev_bg)
                c.border = Border(left=badge_side, right=badge_side,
                                  top=badge_side, bottom=badge_side)
    
            elif col_i == 7:  # Risk Score
                try:
                    risk = float(val)
                    c.value = round(risk, 1)
                    c.number_format = '0.0'
                    risk_color = (C['accent_purple'] if risk >= 80 else
                                  C['accent_red']    if risk >= 60 else
                                  C['accent_amber']  if risk >= 30 else
                                  C['accent_green'])
                    c.font = _font(risk_color, bold=True, size=10)
                except (TypeError, ValueError):
                    c.value = disp
                    c.font  = _font(C['text_primary'], size=9)
                c.alignment = _align('center', 'center')
    
            elif col_i == 8:  # Confidence
                c.value = disp
                c.font  = _font(C['accent_green'], size=9)
                c.alignment = _align('center', 'center')
    
            elif col_i == 9:  # Campaign
                c.value = disp
                if disp.startswith('CAM-'):
                    c.font = _font(C['accent_purple'], bold=True, size=9, name='Consolas')
                else:
                    c.font = _font(C['text_muted'], size=9)
                c.alignment = _align('center', 'center')
    
            elif col_i == 10:  # Escalation
                is_esc = 'tier' in disp.lower() or disp.lower() in ('yes', 'escalated to tier-2')
                c.value = '⬆ Escalated' if is_esc else '— Review'
                c.font  = _font(C['accent_red'] if is_esc else C['text_muted'],
                                bold=is_esc, size=9)
                c.alignment = _align('center', 'center')
    
            elif col_i == 11:  # MITRE
                # Parse JSON string if needed
                mitre_str = disp
                try:
                    import json
                    m = json.loads(disp)
                    tid   = m.get('technique_id', '')
                    tname = m.get('technique_name', '')
                    tact  = m.get('tactic', '')
                    mitre_str = f"{tid} – {tname} ({tact})" if tid else disp
                except Exception:
                    pass
                c.value = mitre_str
                c.font  = _font(C['text_secondary'], size=8, italic=True)
                c.alignment = _align('left', 'center', wrap=True)
    
            elif col_i == 12:  # AI Summary
                c.value = disp
                c.font  = _font(C['text_primary'], size=9)
                c.alignment = _align('left', 'top', wrap=True)
    
            elif col_i == 13:  # Recommendation
                c.value = disp
                c.font  = _font(C['accent_amber'], size=9)
                c.alignment = _align('left', 'top', wrap=True)
    
    # ══════════════════════════════════════════════════════════════════════════════
    # SHEET 2 — EXECUTIVE SUMMARY
    # ══════════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Executive Summary')
    ws2.sheet_view.showGridLines = False
    ws2.sheet_view.showRowColHeaders = False
    
    ws2.column_dimensions['A'].width = 4
    ws2.column_dimensions['B'].width = 28
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 4
    ws2.column_dimensions['E'].width = 28
    ws2.column_dimensions['F'].width = 18
    ws2.column_dimensions['G'].width = 4
    
    for row in range(1, 40):
        ws2.row_dimensions[row].height = 18
        for col in range(1, 8):
            ws2.cell(row, col).fill = _fill(C['bg_dark'])
    
    # Title
    ws2.row_dimensions[1].height = 44
    ws2.merge_cells('A1:G1')
    c = ws2['A1']
    c.value = '  SENTINEL CORE  ·  EXECUTIVE SUMMARY'
    c.font  = _font(C['white'], bold=True, size=16)
    c.fill  = _fill(C['bg_dark'])
    c.alignment = _align('left', 'center')
    c.border = Border(left=Side(border_style='thick', color=C['accent_blue']))
    
    ws2.row_dimensions[2].height = 22
    ws2.merge_cells('A2:G2')
    c = ws2['A2']
    c.value = f"  Generated: {generated_at}  ·  Total Incidents: {len(data_dicts)}"
    c.font  = _font(C['text_secondary'], size=9, italic=True)
    c.fill  = _fill(C['bg_dark'])
    c.alignment = _align('left', 'center')
    
    ws2.row_dimensions[3].height = 14
    
    # Severity breakdown header
    from collections import Counter
    sev_counts  = Counter(d.get('severity') or 'Unknown' for d in data_dicts)
    type_counts = Counter(d.get('alert_type') or 'Unknown' for d in data_dicts)
    total       = len(data_dicts)
    
    def _s2_section_header(ws, row, col_start, col_end, label, color):
        ws.row_dimensions[row].height = 26
        ws.merge_cells(f'{get_column_letter(col_start)}{row}:{get_column_letter(col_end)}{row}')
        c = ws.cell(row, col_start)
        c.value = label
        c.font  = _font(color, bold=True, size=10)
        c.fill  = _fill(C['bg_card'])
        c.alignment = _align('left', 'center')
        c.border = Border(bottom=Side(border_style='medium', color=color))
        for col in range(col_start+1, col_end+1):
            cell = ws.cell(row, col)
            cell.fill   = _fill(C['bg_card'])
            cell.border = Border(bottom=Side(border_style='medium', color=color))
    
    _s2_section_header(ws2, 4, 2, 3, 'SEVERITY BREAKDOWN', C['accent_blue'])
    _s2_section_header(ws2, 4, 5, 6, 'ALERT TYPE DISTRIBUTION', C['accent_blue'])
    
    sev_order = [
        ('Critical', C['accent_purple']),
        ('High',     C['accent_red']),
        ('Medium',   C['accent_amber']),
        ('Low',      C['accent_green']),
    ]
    for i, (sev_label, color) in enumerate(sev_order):
        r     = 5 + i * 3
        count = sev_counts.get(sev_label, 0)
        pct   = count / total if total else 0
        ws2.row_dimensions[r].height   = 14
        ws2.row_dimensions[r+1].height = 26
        ws2.row_dimensions[r+2].height = 8
    
        ws2.merge_cells(f'B{r}:C{r}')
        c = ws2.cell(r, 2); c.value = sev_label.upper()
        c.font = _font(color, bold=True, size=8); c.fill = _fill(C['bg_card'])
        c.alignment = _align('left', 'center'); ws2.cell(r, 3).fill = _fill(C['bg_card'])
    
        c2 = ws2.cell(r+1, 2); c2.value = count
        c2.font = _font(color, bold=True, size=22); c2.fill = _fill(C['bg_card'])
        c2.alignment = _align('left', 'center')
    
        bar = '█' * int(pct * 12) + '░' * (12 - int(pct * 12))
        c3 = ws2.cell(r+1, 3); c3.value = f'{bar}  {int(pct*100)}%'
        c3.font = _font(color, size=9, name='Consolas'); c3.fill = _fill(C['bg_card'])
        c3.alignment = _align('left', 'center')
    
        for col in (2, 3): ws2.cell(r+2, col).fill = _fill(C['bg_dark'])
    
    type_colors = [C['accent_amber'], C['accent_purple'], C['accent_blue'],
                   C['accent_green'], C['accent_red']]
    for i, (atype, count) in enumerate(type_counts.most_common(4)):
        r     = 5 + i * 3
        color = type_colors[i % len(type_colors)]
        pct   = count / total if total else 0
        ws2.row_dimensions[r].height   = 14
        ws2.row_dimensions[r+1].height = 26
        ws2.row_dimensions[r+2].height = 8
    
        ws2.merge_cells(f'E{r}:F{r}')
        c = ws2.cell(r, 5); c.value = atype.upper()
        c.font = _font(color, bold=True, size=8); c.fill = _fill(C['bg_card'])
        c.alignment = _align('left', 'center'); ws2.cell(r, 6).fill = _fill(C['bg_card'])
    
        c2 = ws2.cell(r+1, 5); c2.value = count
        c2.font = _font(color, bold=True, size=22); c2.fill = _fill(C['bg_card'])
        c2.alignment = _align('left', 'center')
    
        bar = '█' * int(pct * 12) + '░' * (12 - int(pct * 12))
        c3 = ws2.cell(r+1, 6); c3.value = f'{bar}  {int(pct*100)}%'
        c3.font = _font(color, size=9, name='Consolas'); c3.fill = _fill(C['bg_card'])
        c3.alignment = _align('left', 'center')
    
        for col in (5, 6): ws2.cell(r+2, col).fill = _fill(C['bg_dark'])
    
    # Key stats row
    stat_row = 22
    ws2.row_dimensions[stat_row].height   = 26
    ws2.row_dimensions[stat_row+1].height = 18
    ws2.row_dimensions[stat_row+2].height = 30
    ws2.row_dimensions[stat_row+3].height = 8
    
    ws2.merge_cells(f'B{stat_row}:F{stat_row}')
    c = ws2.cell(stat_row, 2); c.value = 'KEY STATISTICS'
    c.font = _font(C['accent_blue'], bold=True, size=10); c.fill = _fill(C['bg_card'])
    c.alignment = _align('left', 'center')
    c.border = Border(bottom=Side(border_style='medium', color=C['accent_blue']))
    for col in range(3, 7):
        cell = ws2.cell(stat_row, col)
        cell.fill   = _fill(C['bg_card'])
        cell.border = Border(bottom=Side(border_style='medium', color=C['accent_blue']))
    
    escalated = sum(1 for d in data_dicts
                    if 'tier' in str(d.get('escalation') or '').lower()
                    or str(d.get('escalation') or '').lower() == 'yes')
    campaigns = len(set(d.get('campaign_id') for d in data_dicts
                        if d.get('campaign_id') and d.get('campaign_id') != 'standalone'))
    resolved  = sum(1 for d in data_dicts if str(d.get('status') or '').lower() == 'resolved')
    mitigated = f"{resolved / total * 100:.1f}%" if total else "0.0%"
    
    stats = [
        ('TOTAL INCIDENTS', total,      C['accent_blue']),
        ('ESCALATED',       escalated,  C['accent_red']),
        ('CAMPAIGNS',       campaigns,  C['accent_purple']),
        ('MITIGATED',       mitigated,  C['accent_amber']),
    ]
    stat_cols = [2, 3, 5, 6]
    for i, (label, value, color) in enumerate(stats):
        col = stat_cols[i]
        c = ws2.cell(stat_row+1, col); c.value = label
        c.font = _font(C['text_secondary'], bold=True, size=8); c.fill = _fill(C['bg_card'])
        c.alignment = _align('center', 'center')
    
        c2 = ws2.cell(stat_row+2, col); c2.value = value
        c2.font = _font(color, bold=True, size=22); c2.fill = _fill(C['bg_card'])
        c2.alignment = _align('center', 'center')
    
    # ══════════════════════════════════════════════════════════════════════════════
    # SHEET 3 — TOP THREATS (Critical & High only)
    # ══════════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('Top Threats')
    ws3.sheet_view.showGridLines = False
    ws3.sheet_view.showRowColHeaders = False
    
    top_cols = [20, 18, 16, 14, 10, 32, 52]
    for i, w in enumerate(top_cols, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    
    ws3.row_dimensions[1].height = 44
    ws3.merge_cells('A1:G1')
    c = ws3['A1']
    c.value = '  SENTINEL CORE  ·  CRITICAL & HIGH PRIORITY THREATS'
    c.font  = _font(C['white'], bold=True, size=14)
    c.fill  = _fill(C['bg_dark'])
    c.alignment = _align('left', 'center')
    c.border = Border(left=Side(border_style='thick', color=C['accent_red']))
    
    ws3.row_dimensions[2].height = 28
    top_headers = ['INCIDENT ID','TIMESTAMP','SOURCE IP','CAMPAIGN',
                   'RISK SCORE','MITRE ATT&CK','AI SUMMARY  →  RECOMMENDATION']
    for i, h in enumerate(top_headers, 1):
        c = ws3.cell(2, i); c.value = h
        c.font = _font(C['accent_red'], bold=True, size=9)
        c.fill = _fill(C['bg_header']); c.alignment = _align('center', 'center')
        c.border = Border(bottom=Side(border_style='medium', color=C['accent_red']))
    
    ws3.freeze_panes = 'A3'
    
    critical_rows = [d for d in data_dicts if str(d.get('severity')) in ('Critical', 'High')]
    for row_i, d in enumerate(critical_rows):
        r  = row_i + 3
        bg = C['bg_card'] if row_i % 2 == 0 else C['bg_alt']
        ws3.row_dimensions[r].height = 55
        bb = _border_bottom()
    
        # MITRE parse
        mitre_raw = str(d.get('mitre_technique') or '')
        try:
            import json as _json
            m = _json.loads(mitre_raw)
            mitre_disp = f"{m.get('technique_id','')} – {m.get('technique_name','')} ({m.get('tactic','')})"
        except Exception:
            mitre_disp = mitre_raw
    
        combined = f"{d.get('incident_summary','')}\n\n→  {d.get('recommended_action','')}"
    
        row_cells = [
            (str(d.get('incident_id') or ''),   'Consolas', 9,  True,  C['accent_blue'],   'left',   False),
            (str(d.get('timestamp') or ''),      'Consolas', 9,  False, C['text_secondary'],'center', False),
            (str(d.get('source_ip') or ''),      'Consolas', 9,  False, C['text_primary'],  'center', False),
            (str(d.get('campaign_id') or ''),    'Consolas', 9,  True,  C['accent_purple'], 'center', False),
            (str(d.get('risk_score') or ''),     'Calibri',  11, True,  C['accent_purple'], 'center', False),
            (mitre_disp,                         'Calibri',  8,  False, C['text_secondary'],'left',   True),
            (combined,                           'Calibri',  9,  False, C['text_primary'],  'left',   True),
        ]
        for col_i, (val, fname, fsize, fbold, fcolor, halign, fwrap) in enumerate(row_cells, 1):
            c = ws3.cell(r, col_i)
            c.value = val; c.fill = _fill(bg); c.border = bb
            c.font  = Font(name=fname, size=fsize, bold=fbold, color=fcolor)
            c.alignment = _align(halign, 'top', wrap=fwrap)
    
    # ── Serialize ──────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()
    
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=soc-ai-incident-report.xlsx"},
    )


@router.get("/report/export/csv")
def export_report_csv():
    """Returns an enriched, SOC-ready CSV report of all processed alerts."""
    import re
    from datetime import datetime, timedelta
    from collections import Counter

    try:
        with engine.connect() as conn:
            rows = conn.execute(text("SELECT * FROM alerts")).fetchall()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    if not rows:
        raise HTTPException(status_code=404, detail="No alerts found. Run Analysis first.")

    raw_data = [dict(r._mapping) for r in rows]

    # --- 1. Aggregation Sweep ---
    ip_counts = Counter(d.get("source_ip") for d in raw_data if d.get("source_ip"))
    type_counts = Counter(d.get("alert_type") for d in raw_data if d.get("alert_type"))

    # --- 2. Deduplication (5-minute window) ---
    def get_dedup_key(d):
        ts = d.get("timestamp")
        # Quantize timestamp to 5-minute buckets for grouping rapid fire alerts
        if ts:
            try:
                # Assuming ISO format or similar
                dt = datetime.fromisoformat(str(ts).replace("Z", "+00:00"))
                ts_bucket = dt.replace(second=0, microsecond=0, minute=(dt.minute // 5) * 5)
            except:
                ts_bucket = str(ts)[:16]  # Fallback to string slice YYYY-MM-DD HH:MM
        else:
            ts_bucket = "unknown"
        
        return (d.get("source_ip"), d.get("destination_ip"), d.get("alert_type"), d.get("port"), ts_bucket)

    seen = {}
    deduplicated = []
    for d in raw_data:
        key = get_dedup_key(d)
        if key not in seen:
            seen[key] = True
            deduplicated.append(d)

    # --- 3. Enrichment Helper Logic ---
    def get_threat_category(alert_type):
        at = str(alert_type).lower()
        if any(x in at for x in ["brute", "password", "login"]): return "Credential Access / Brute Force"
        if any(x in at for x in ["malware", "virus", "trojan", "ransomware"]): return "Malware Execution"
        if any(x in at for x in ["scan", "recon", "nmap", "probe"]): return "Reconnaissance"
        if any(x in at for x in ["exfil", "upload", "leak"]): return "Data Exfiltration"
        if any(x in at for x in ["ddos", "flood", "dos"]): return "Denial of Service"
        if any(x in at for x in ["sql", "inject", "exploit", "rce"]): return "Exploitation"
        return "Suspicious Activity"

    def get_geo(ip):
        ip_str = str(ip)
        if ip_str.startswith(("10.", "192.168.", "172.")):
            if ip_str.startswith("172."):
                # Check 172.16.0.0 – 172.31.255.255
                m = re.match(r"172\.(\d+)\.", ip_str)
                if m and 16 <= int(m.group(1)) <= 31: return "Internal Network"
            else:
                return "Internal Network"
        return "External (WAN / Public)"

    # --- 4. Final Processing Pass ---
    enriched = []
    for d in deduplicated:
        risk = d.get("risk_score", 0)
        try: risk = float(risk)
        except: risk = 0.0

        src_ip = d.get("source_ip", "")
        offenses = ip_counts.get(src_ip, 0)

        # Build Enriched Object
        row = {
            "incident_id": d.get("incident_id", ""),
            "std_timestamp": str(d.get("timestamp", ""))[:19],
            "source_ip": src_ip,
            "destination_ip": d.get("destination_ip", ""),
            "port": d.get("port", ""),
            "protocol": d.get("protocol", ""),
            "alert_type": d.get("alert_type", ""),
            "severity_level": d.get("severity", "Medium"),
            "threat_category": get_threat_category(d.get("alert_type")),
            "risk_score": round(risk, 1),
            "confidence_score": d.get("confidence", ""),
            "campaign_id": d.get("campaign_id", "standalone"),
            "escalation_status": d.get("escalation", "No"),
            "incident_status": d.get("status", "New"),
            "anomaly_flag": "Yes" if (risk >= 75 or offenses > 10) else "No",
            "src_ip_reputation": "Malicious" if offenses > 5 else "Suspicious" if offenses > 2 else "Safe",
            "geo_location": get_geo(src_ip),
            "repeat_ip_offenses": offenses,
            "alert_count_per_type": type_counts.get(d.get("alert_type"), 0),
            "alert_summary": str(d.get("incident_summary", "")).replace("\n", " ").split(". ")[0],
            "recommended_action": str(d.get("recommended_action", "")).replace("\n", " ")
        }
        enriched.append(row)

    # Sort by risk_score DESC
    enriched.sort(key=lambda x: x["risk_score"], reverse=True)

    # --- 5. Format CSV ---
    if not enriched:
        headers = [] # Should not happen given rows check
    else:
        headers = list(enriched[0].keys())

    def escape(val):
        s = str(val) if val is not None else ""
        return '"' + s.replace('"', '""').replace("\r", "").replace("\n", " ") + '"'

    output_lines = [",".join(headers)]
    for row in enriched:
        output_lines.append(",".join(escape(row.get(h, "")) for h in headers))

    filename = f"sentinel_soc_analytics_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.csv"
    
    return Response(
        content="\n".join(output_lines),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/report/{incident_id}")
def get_incident_report(incident_id: str):
    """Returns a structured AI-generated JSON report for one incident."""
    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT * FROM alerts WHERE incident_id = :id"),
            {"id": incident_id},
        ).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Incident not found")

    data = dict(row._mapping)
    return {
        "incident_id":    data.get("incident_id"),
        "summary":        data.get("incident_summary"),
        "recommendation": data.get("recommended_action"),
        "severity":       data.get("severity"),
        "risk_score":     data.get("risk_score"),
        "alert_type":     data.get("alert_type"),
        "source_ip":      data.get("source_ip"),
        "destination_ip": data.get("destination_ip"),
        "timestamp":      str(data.get("timestamp", "")),
        "mitre":          data.get("mitre_technique"),
        "soc_playbook":   data.get("soc_playbook_action"),
        "automation":     data.get("automation_result"),
        "escalation":     data.get("escalation"),
        "campaign_id":    data.get("campaign_id"),
        "notes":          data.get("notes", ""),
    }
